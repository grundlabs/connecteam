# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
import sys
import os

# Configure stdout encoding for Windows compatibility
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

def process_timesheet(input_file, output_file=None):
    """
    Process employee timesheet shift report
    
    New format: Each employee has name on one row, followed by their shifts
    """
    
    # Read Excel file
    try:
        df = pd.read_excel(input_file, sheet_name=0)
    except Exception as e:
        print(f"ERROR: Could not read file: {e}")
        return None
    
    print(f"Loaded file with {len(df)} rows")
    print(f"Columns: {df.columns.tolist()}\n")
    
    # Define allowed shift types
    ALLOWED_TYPES = ['Hosszú', 'Leo', 'Winston', 'Mogumba', 'Konyha', 'Nappalos', 'Poharas', 'Kávézó pult', 'Rács']
    
    # Collect all unique dates from the file
    unique_dates = []
    seen_dates = set()
    
    for val in df['Start Date']:
        if pd.notna(val):
            date_str = str(val).split()[0]  # Remove time portion (e.g., "2026-03-13 00:00:00" -> "2026-03-13")
            if date_str not in seen_dates:
                seen_dates.add(date_str)
                unique_dates.append(val)
    
    print(f"Unique dates found: {unique_dates}")
    
    # Error handling: Check if more than 2 unique dates
    if len(unique_dates) > 2:
        print("\nERROR: Hibás bemeneti fájl, több mint 1 napot tartalmaz.")
        return None
    
    if len(unique_dates) == 0:
        print("\nERROR: No valid dates found")
        return None
    
    # Choose chronologically earlier date
    if len(unique_dates) == 2:
        try:
            dates_parsed = []
            for d in unique_dates:
                if isinstance(d, str):
                    date_clean = d.split()[0]
                    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
                        try:
                            dates_parsed.append((d, datetime.strptime(date_clean, fmt)))
                            break
                        except:
                            continue
                else:
                    dates_parsed.append((d, d))
            
            dates_parsed.sort(key=lambda x: x[1])
            filter_date = dates_parsed[0][0]
            print(f"Two dates found, using chronologically earlier: {filter_date}")
        except:
            filter_date = unique_dates[0]
            print(f"Using first date: {filter_date}")
    else:
        filter_date = unique_dates[0]
        print(f"Single date found: {filter_date}")
    
    print(f"\nFiltering for date: {filter_date}\n")
    
    # Process records
    records = []
    current_employee = None
    employee_shifts = []
    
    stats = {
        'employees_processed': 0,
        'shifts_found': 0,
        'filtered_by_type': 0,
        'filtered_by_date': 0,
        'filtered_by_name': 0,
        'filtered_by_special_case': 0,
        'output_records': 0
    }
    
    for idx, row in df.iterrows():
        # Check if this is an employee name row (has First name and Last name)
        if pd.notna(row['First name']) and pd.notna(row['Last name']):
            # Save previous employee's shift if any
            if current_employee and employee_shifts:
                # Find chronologically earlier shift for this employee
                selected_shift = select_shift(employee_shifts, filter_date, ALLOWED_TYPES, current_employee, stats)
                if selected_shift:
                    # Apply name filters
                    first_name, last_name = current_employee
                    
                    # Filter: Exclude Panácz
                    if last_name.lower() == 'panácz':
                        stats['filtered_by_name'] += 1
                    else:
                        # Special case: Bence Horváth in Konyha -> HORVÁTH BENCE POHARAS
                        if (first_name.lower() == 'bence' and 
                            last_name.lower() == 'horváth' and 
                            selected_shift['type'] == 'Konyha'):
                            name = "HORVÁTH BENCE POHARAS"
                        else:
                            name = f"{last_name.upper()} {first_name.upper()}"
                        
                        records.append({
                            'name': name,
                            'in': selected_shift['in'],
                            'out': selected_shift['out']
                        })
                        stats['output_records'] += 1
            
            # Start new employee
            current_employee = (row['First name'], row['Last name'])
            employee_shifts = []
            stats['employees_processed'] += 1
        
        # Check if this is a shift row (has Shift Number)
        elif pd.notna(row.get('Shift Number')):
            shift_type = row.get('Type')
            start_date = row.get('Start Date')
            in_time = row.get('In')
            out_time = row.get('Out')
            
            stats['shifts_found'] += 1
            
            employee_shifts.append({
                'type': shift_type,
                'start_date': start_date,
                'in': in_time,
                'out': out_time,
                'shift_number': row['Shift Number']
            })
    
    # Process last employee
    if current_employee and employee_shifts:
        selected_shift = select_shift(employee_shifts, filter_date, ALLOWED_TYPES, current_employee, stats)
        if selected_shift:
            first_name, last_name = current_employee
            
            if last_name.lower() == 'panácz':
                stats['filtered_by_name'] += 1
            else:
                if (first_name.lower() == 'bence' and 
                    last_name.lower() == 'horváth' and 
                    selected_shift['type'] == 'Konyha'):
                    name = "HORVÁTH BENCE POHARAS"
                else:
                    name = f"{last_name.upper()} {first_name.upper()}"
                
                records.append({
                    'name': name,
                    'in': selected_shift['in'],
                    'out': selected_shift['out']
                })
                stats['output_records'] += 1
    
    # Print statistics
    print(f"Processing summary:")
    print(f"  - Employees processed: {stats['employees_processed']}")
    print(f"  - Total shifts found: {stats['shifts_found']}")
    print(f"  - Filtered by type: {stats['filtered_by_type']}")
    print(f"  - Filtered by date: {stats['filtered_by_date']}")
    print(f"  - Filtered by name (Panácz): {stats['filtered_by_name']}")
    print(f"  - Filtered by special case (Kávézó): {stats['filtered_by_special_case']}")
    print(f"  - Output records: {stats['output_records']}")
    
    if not records:
        print("\nNo valid records found!")
        return None
    
    # Create Excel output
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Records"
    
    # Headers
    headers = ['Name', 'In', 'Out']
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for record in records:
        ws.append([record['name'], str(record['in']), str(record['out'])])
        
        # Center align time columns
        row_num = ws.max_row
        ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    
    # Determine output filename
    if output_file is None:
        base_name = os.path.splitext(os.path.basename(input_file))[0]
        output_file = f'/mnt/user-data/outputs/{base_name}_formatted.xlsx'
    
    wb.save(output_file)
    
    print(f"\n[SUCCESS] Successfully processed {len(records)} records")
    print(f"[SUCCESS] Output saved to: {output_file}")
    print(f"\nSample records:")
    for i in range(min(5, len(records))):
        rec = records[i]
        print(f"  {rec['name']}: {rec['in']} -> {rec['out']}")
    
    return output_file

def select_shift(shifts, filter_date, allowed_types, employee, stats):
    """
    Select the appropriate shift for an employee
    - Filter by date
    - Filter by type (with special case for István Prihoda + Kávézó)
    - If multiple shifts remain, choose chronologically earlier one
    - If only one shift, use it
    """
    
    first_name, last_name = employee
    
    # Filter by date
    valid_shifts = []
    for shift in shifts:
        if pd.notna(shift['start_date']):
            # Compare dates (handle both string and datetime)
            shift_date_str = str(shift['start_date']).split()[0]
            filter_date_str = str(filter_date).split()[0]
            
            if shift_date_str == filter_date_str:
                valid_shifts.append(shift)
            else:
                stats['filtered_by_date'] += 1
        else:
            stats['filtered_by_date'] += 1
    
    if not valid_shifts:
        return None
    
    # Filter by type
    type_filtered = []
    for shift in valid_shifts:
        shift_type = shift['type']
        
        # Special case: Kávézó only allowed for István Prihoda
        if pd.notna(shift_type) and str(shift_type) == 'Kávézó':
            if first_name.lower() == 'istván' and last_name.lower() == 'prihoda':
                type_filtered.append(shift)
            else:
                stats['filtered_by_special_case'] += 1
        elif pd.notna(shift_type) and str(shift_type) in allowed_types:
            type_filtered.append(shift)
        else:
            stats['filtered_by_type'] += 1
    
    if not type_filtered:
        return None
    
    # If only one shift, return it
    if len(type_filtered) == 1:
        return type_filtered[0]
    
    # Multiple shifts: choose chronologically earlier one based on Start Date
    try:
        type_filtered.sort(key=lambda x: x['start_date'])
        return type_filtered[0]
    except:
        # If sorting fails, return first
        return type_filtered[0]

def main():
    if len(sys.argv) < 2:
        print("Usage: python process_timesheets.py <input_file.xlsx> [output_file.xlsx]")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    if not os.path.exists(input_file):
        print(f"ERROR: File not found: {input_file}")
        sys.exit(1)
    
    process_timesheet(input_file, output_file)

if __name__ == "__main__":
    main()
